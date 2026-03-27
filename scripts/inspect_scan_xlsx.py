#!/usr/bin/env python3
"""
Print data rows from a Watchlist strategy scan .xlsx (headers row 4, data from row 5).

Usage:
  python scripts/inspect_scan_xlsx.py
  python scripts/inspect_scan_xlsx.py "D:/path/to/strategy_04_Qtum_scan_20260321_1234.xlsx"

With no path, uses the newest strategy_*_scan_*.xlsx under project output/.
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src.excel_writer import DATA_COLUMNS


def main() -> int:
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
    else:
        out = ROOT / "output"
        files = sorted(out.glob("strategy_*_scan_*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)
        p = files[0] if files else None

    if p is None or not p.is_file():
        print("No spreadsheet found. Pass the full path to your .xlsx or run a scan so output/ contains strategy_*_scan_*.xlsx")
        return 1

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        print("File:", p.resolve())
        print("Sheet:", ws.title)
        hdr = [ws.cell(4, c).value for c in range(1, len(DATA_COLUMNS) + 1)]
        print("Row 4 headers:", hdr)
        n = ws.max_row or 4
        if n < 5:
            print("No data rows (max_row < 5).")
            return 0
        print(f"Data rows 5..{min(n, 20)} (values, data_only=True):")
        any_number = False
        for r in range(5, min(n, 20) + 1):
            row = [ws.cell(r, c).value for c in range(1, len(DATA_COLUMNS) + 1)]
            sym = row[0] if row else None
            nums = [x for x in row[1:] if isinstance(x, (int, float)) and x == x]
            if nums:
                any_number = True
            print(f"  {r}: {row}")
        if not any_number and n >= 5:
            print(
                "\nNote: No numeric cells in sampled rows - metrics were likely empty when saved "
                "(same as live preview dashes)."
            )
    finally:
        wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
