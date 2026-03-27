#!/usr/bin/env python3
"""
Convert existing Excel scan results to HTML reports.
Usage: py convert_xlsx_to_html.py <file.xlsx> [file2.xlsx ...]
       py convert_xlsx_to_html.py output/*.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

# Add project root
sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.excel_writer import DATA_COLUMNS, get_data_rows
from src.html_writer import write_html_report


def convert_one(xlsx_path: Path) -> Path | None:
    """Convert one xlsx file to HTML. Returns HTML path or None."""
    if not xlsx_path.exists():
        print(f"Not found: {xlsx_path}")
        return None
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb.active
    # Read metadata from row 2
    strategy_url = ws.cell(row=2, column=1).value or ""
    original_pair = ws.cell(row=2, column=2).value or "Unknown"
    wb.close()

    # Extract strategy index and name from filename (strategy_01_CAKEUSDT_P_scan_20260318_2015.xlsx)
    stem = xlsx_path.stem
    parts = stem.split("_")
    strat_idx = 1
    if len(parts) >= 2 and parts[0] == "strategy" and parts[1].isdigit():
        strat_idx = int(parts[1])

    rows = get_data_rows(ws)
    if not rows:
        print(f"No data rows in {xlsx_path.name}")
        return None

    html_path = write_html_report(
        output_dir=xlsx_path.parent,
        strategy_index=strat_idx,
        strategy_url=strategy_url,
        original_pair=original_pair,
        strategy_name=None,
        rows=rows,
        xlsx_path=xlsx_path,
    )
    print(f"  -> {html_path.name}")
    return html_path


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: py convert_xlsx_to_html.py <file.xlsx> [file2.xlsx ...]")
        print("       py convert_xlsx_to_html.py output/*.xlsx")
        sys.exit(1)

    paths = []
    for arg in sys.argv[1:]:
        p = Path(arg)
        if p.is_file() and p.suffix.lower() == ".xlsx":
            paths.append(p)
        elif p.is_dir():
            paths.extend(p.glob("*.xlsx"))
        else:
            paths.extend(Path(".").glob(arg))

    paths = sorted(set(paths))
    if not paths:
        print("No .xlsx files found")
        sys.exit(1)

    print(f"Converting {len(paths)} file(s)...")
    for xlsx_path in paths:
        convert_one(xlsx_path)
    print("Done.")


if __name__ == "__main__":
    main()
