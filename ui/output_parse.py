"""
Parse Watchlist Scanner output filenames into structured fields for the UI.

Pattern (from excel_writer / html_writer):
  strategy_{NN}_{slug}_{scan|deep_scan}_{YYYYMMDD}_{HHMM}.xlsx|.html
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional


# Slug is non-greedy so names like ..._QTUMUSDT_P_deep_scan_... parse correctly.
_OUTPUT_RE = re.compile(
    r"^strategy_(\d+)_(.+?)_((?:deep_scan)|(?:scan))_(\d{8})_(\d{4})\.(xlsx|html)$",
    re.IGNORECASE,
)


@dataclass
class ParsedOutput:
    strategy_num: str
    slug: str
    pair_display: str
    scan_kind: str  # "Deep scan" | "Pass 1 scan"
    run_at: Optional[datetime]
    run_at_label: str
    extension: str  # xlsx | html
    format_label: str  # Spreadsheet | Report


def slug_to_pair_display(slug: str) -> str:
    """Best-effort: QTUMUSDT_P -> QTUMUSDT.P (perpetual suffix)."""
    if not slug:
        return "—"
    s = slug.strip()
    if s.endswith("_P") and len(s) > 2:
        return s[:-2] + ".P"
    return s.replace("_", " ")


def parse_output_path(path: Path) -> Optional[ParsedOutput]:
    m = _OUTPUT_RE.match(path.name)
    if not m:
        return None
    num, slug, kind, ymd, hm, ext = m.groups()
    ext = ext.lower()
    scan_kind = "Deep scan" if kind.lower() == "deep_scan" else "Pass 1 scan"
    run_at = None
    run_at_label = f"{ymd[:4]}-{ymd[4:6]}-{ymd[6:8]} {hm[:2]}:{hm[2:4]}"
    try:
        run_at = datetime.strptime(f"{ymd}_{hm}", "%Y%m%d_%H%M")
        run_at_label = run_at.strftime("%d %b %Y · %H:%M")
    except ValueError:
        pass
    fmt = "Spreadsheet" if ext == "xlsx" else "HTML report"
    return ParsedOutput(
        strategy_num=num,
        slug=slug,
        pair_display=slug_to_pair_display(slug),
        scan_kind=scan_kind,
        run_at=run_at,
        run_at_label=run_at_label,
        extension=ext,
        format_label=fmt,
    )


def queue_job_label(num: str) -> str:
    """Queue row id embedded in filenames as strategy_{NN}_... — not the order of strategies in config.yaml."""
    return f"Job {int(num):02d}" if num.isdigit() else f"Job {num}"


def strategy_label(num: str) -> str:
    """Alias of queue_job_label (legacy name)."""
    return queue_job_label(num)


def _count_xlsx_data_rows(path: Path) -> Optional[int]:
    """Count data rows (excel_writer: headers row 4, data from row 5)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            ws = wb.active
            n = 0
            for row in ws.iter_rows(min_row=5, min_col=1, max_col=1, values_only=True):
                v = row[0] if row else None
                if v is not None and str(v).strip():
                    n += 1
            return n
        finally:
            wb.close()
    except Exception:
        return None


def _count_html_tbody_rows(path: Path) -> Optional[int]:
    """Count <tr> inside <tbody> (one row per pair in html_writer reports)."""
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        low = text.lower()
        i = low.find("<tbody")
        if i < 0:
            return 0
        j = low.find("</tbody>", i)
        if j < 0:
            return None
        chunk = text[i:j]
        return chunk.lower().count("<tr")
    except Exception:
        return None


def count_pairs_in_output_file(path: Path) -> Optional[int]:
    """
    Number of pair/result rows in a scan output file.
    None if the file could not be read.
    """
    if not path.exists():
        return None
    suf = path.suffix.lower()
    if suf == ".xlsx":
        return _count_xlsx_data_rows(path)
    if suf in (".html", ".htm"):
        return _count_html_tbody_rows(path)
    return None
