"""
Excel writer for strategy scan results. One file per strategy, metadata header row, save after each pair.
"""
from __future__ import annotations

import os
import shutil
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Column order for data rows (matches scraper METRIC_KEYS)
DATA_COLUMNS = [
    "Symbol",
    "Net Profit",
    "Net Profit %",
    "Gross Profit",
    "Gross Loss",
    "Max Drawdown",
    "Max Drawdown %",
    "Sharpe Ratio",
    "Sortino Ratio",
    "Win Rate %",
    "# Trades",
    "Profit Factor",
]

# Map display column names to scraper metric keys
COL_TO_METRIC_KEY = {
    "Net Profit": "net_profit",
    "Net Profit %": "net_profit_pct",
    "Gross Profit": "gross_profit",
    "Gross Loss": "gross_loss",
    "Max Drawdown": "max_drawdown",
    "Max Drawdown %": "max_drawdown_pct",
    "Sharpe Ratio": "sharpe_ratio",
    "Sortino Ratio": "sortino_ratio",
    "Win Rate %": "win_rate_pct",
    "# Trades": "total_trades",
    "Profit Factor": "profit_factor",
}


def _slug_for_filename(pair: str) -> str:
    """Convert pair to filename-safe slug, e.g. BYBIT:BBUSDT.P -> BBUSDT_P."""
    if not pair or not pair.strip():
        return ""
    s = pair.strip().replace(":", "_").replace(".", "_")
    return s if s else ""


def create_workbook(
    output_dir: Path,
    strategy_index: int,
    strategy_url: str,
    original_pair: str,
    strategy_name=None,
    output_suffix: str = "scan",
) -> tuple[Workbook, Path]:
    """Create a new workbook with metadata row, return workbook and file path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Strategy Results"

    # Row 1: Metadata
    ws["A1"] = "Strategy Link"
    ws["B1"] = "Original Pair"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    # Hyperlink for strategy URL
    ws["A2"] = strategy_url
    ws["A2"].hyperlink = strategy_url
    ws["A2"].font = Font(color="0563C1", underline="single")
    ws["B2"] = original_pair

    # Row 4: Data column headers
    for col_idx, col_name in enumerate(DATA_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Filename: prefer original pair (e.g. BBUSDT_P) when available, else strategy name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    pair_slug = _slug_for_filename(original_pair) if original_pair and original_pair != "Unknown" else ""
    slug = pair_slug or (strategy_name or f"strategy_{strategy_index:02d}").strip().replace(" ", "_")
    filename = f"strategy_{strategy_index:02d}_{slug}_{output_suffix}_{timestamp}.xlsx"
    filepath = output_dir / filename

    return wb, filepath


def row_values_for_metrics(symbol: str, metrics: Optional[dict[str, Any]]) -> list[Any]:
    """Build one data row (len == len(DATA_COLUMNS)) for symbol + scraper metrics dict."""
    m = metrics or {}
    values: list[Any] = [symbol]
    for col in DATA_COLUMNS[1:]:
        key = COL_TO_METRIC_KEY.get(col, col.replace(" ", "_").lower())
        if key in m:
            val = m[key]
        elif col in m:
            val = m[col]
        else:
            val = None
        if val is None:
            val = ""
        elif isinstance(val, (int, float)):
            val = val if val == val else ""  # NaN check
        values.append(val)
    return values


def append_result_row(ws, row_num: int, symbol: str, metrics: dict[str, Any]) -> None:
    """Append one result row."""
    values = row_values_for_metrics(symbol, metrics)
    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=row_num, column=col_idx, value=val)


def get_next_data_row(ws) -> int:
    """Get the next row number for data (headers at row 4, data starts row 5)."""
    return ws.max_row + 1 if ws.max_row >= 4 else 5


def sort_data_by_net_profit(ws) -> None:
    """Sort data rows (5+) by Net Profit (column B) descending. Highest first."""
    if ws.max_row < 5:
        return
    num_cols = len(DATA_COLUMNS) + 1  # Symbol + metric columns
    rows = []
    for r in range(5, ws.max_row + 1):
        row_data = [ws.cell(row=r, column=c).value for c in range(1, num_cols + 1)]
        rows.append(row_data)
    net_profit_col = 2  # Column B

    def _key(row):
        val = row[net_profit_col - 1] if len(row) >= net_profit_col else None
        if val is None or val == "":
            return float("-inf")
        try:
            return -float(val) if isinstance(val, (int, float)) else float("-inf")
        except (TypeError, ValueError):
            return float("-inf")

    rows.sort(key=_key)
    for i, row_data in enumerate(rows):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=5 + i, column=c, value=val)


def _canonical_symbol(symbol: str) -> str:
    """Normalize symbol for dedup: BYBIT:BTCUSDT.P -> BTCUSDT."""
    s = (symbol or "").strip().upper().replace(" ", "")
    if ":" in s:
        s = s.split(":", 1)[1]
    if s.endswith(".P"):
        s = s[:-2]
    return s


def get_completed_symbols(ws) -> set[str]:
    """Return set of canonical symbols already in data rows (5+). For resume support."""
    out = set()
    if ws.max_row < 5:
        return out
    for r in range(5, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str):
            out.add(_canonical_symbol(val))
    return out


def read_pass1_workbook_progress(path: Path) -> tuple[set[str], list[str]]:
    """
    Read symbol column (A, row 5+) from a Pass 1 workbook.
    Returns (canonical_done_set, symbols_in_row_order) for resume / disk recovery.
    """
    path = Path(path)
    if not path.is_file():
        return set(), []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        ordered: list[str] = []
        done: set[str] = set()
        if ws.max_row < 5:
            return done, ordered
        for r in range(5, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val and isinstance(val, str):
                s = val.strip()
                if s:
                    ordered.append(s)
                    done.add(_canonical_symbol(s))
        return done, ordered
    finally:
        wb.close()


def open_workbook_for_append(filepath: Path) -> tuple[Workbook, Path]:
    """Load existing workbook for appending rows. Returns (workbook, filepath)."""
    wb = load_workbook(filepath, read_only=False, data_only=False)
    return wb, filepath


def get_data_rows(ws) -> list[list]:
    """Return data rows (5+) from worksheet as list of lists. Call after sort_data_by_net_profit."""
    if ws.max_row < 5:
        return []
    num_cols = len(DATA_COLUMNS)
    rows = []
    for r in range(5, ws.max_row + 1):
        row_data = [ws.cell(row=r, column=c).value for c in range(1, num_cols + 1)]
        rows.append(row_data)
    return rows


def read_scan_preview_rows(filepath: Path, *, max_rows: int = 50) -> list[list[Any]]:
    """
    Read data rows from a scan workbook between worker saves.

    Copies the file to a temp path before opening so we never read a half-written
    .xlsx while the worker thread is inside ``save()`` (common on Windows).

    Uses ``data_only=False`` so stored literals are always returned (matches Excel grid).

    Rows are sorted by Net Profit % (column index 2) descending, like the final HTML table.
    Returns [] if the file is missing, locked, or unreadable.
    """
    filepath = Path(filepath).resolve()
    if not filepath.is_file():
        return []
    num_cols = len(DATA_COLUMNS)

    def _extract(ws) -> list[list[Any]]:
        out: list[list[Any]] = []
        max_r = getattr(ws, "max_row", None) or 0
        if max_r < 5:
            return out
        for r in range(5, max_r + 1):
            row_data = [ws.cell(row=r, column=c).value for c in range(1, num_cols + 1)]
            first = row_data[0]
            if first is None or (isinstance(first, str) and not str(first).strip()):
                continue
            out.append(row_data)
        return out

    def _load_rows(path: Path) -> list[list[Any]]:
        wb = load_workbook(path, read_only=False, data_only=False, keep_vba=False)
        try:
            return _extract(wb.active)
        finally:
            wb.close()

    raw: list[list[Any]] = []
    tmp: Path | None = None
    try:
        fd, tmp_name = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        tmp = Path(tmp_name)
        for attempt in range(6):
            try:
                shutil.copy2(filepath, tmp)
                raw = _load_rows(tmp)
                break
            except Exception:
                time.sleep(0.05 * (attempt + 1))
        if not raw:
            try:
                raw = _load_rows(filepath)
            except Exception:
                raw = []
    finally:
        if tmp is not None:
            try:
                tmp.unlink(missing_ok=True)
            except OSError:
                pass

    if not raw:
        return []

    def _np_pct_key(r: list[Any]) -> float:
        try:
            v = r[2] if len(r) > 2 else None
            if v is None or v == "":
                return float("inf")
            return -float(v)
        except (TypeError, ValueError):
            return float("inf")

    raw.sort(key=_np_pct_key)
    return raw[:max_rows]


def sort_preview_rows_by_net_pct(rows: list[list[Any]]) -> list[list[Any]]:
    """Sort copy of preview rows by Net Profit % (column index 2) descending."""

    def _np_pct_key(r: list[Any]) -> float:
        try:
            v = r[2] if len(r) > 2 else None
            if v is None or v == "":
                return float("inf")
            return -float(v)
        except (TypeError, ValueError):
            return float("inf")

    return sorted(rows, key=_np_pct_key)
